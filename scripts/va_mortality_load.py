#!/usr/bin/env python3
"""Deterministic (LLM-free) loader for the VA veteran mortality appendices.

Rebuilds the two dead VA sources from the 2026-08-09 triage (Chris: Option 1):
  - fed_va_suicide_appendix was a flattened-PDF scrape (positional COL_n junk);
    the official xlsx appendices replace it as TWO real sources:
      fed_va_suicide_national  (cohort x year national suicide counts/rates)
      fed_va_suicide_state     (state-level suicides by state/sex/age/method)
  - fed_va_allcause_mortality was a broken PDF list-of-figures scrape; the
    2018-2023 xlsx appendix replaces it under the SAME source_id.

All three xlsx URLs live-verified 2026-08-09 (HTTP 200, real xlsx). Shapes:
  suicide files: 2 junk banner rows, header row 3, tidy rows below.
  all-cause file: per sheet, SIX 8-column year blocks side by side
  (year stamp above each block) -> unpivoted to long (cohort, year, rank, cause).

    python scripts/va_mortality_load.py            # preview (downloads + parses only)
    python scripts/va_mortality_load.py --run      # land + quality gate + register
"""
from __future__ import annotations

import argparse
import hashlib
import io
import re
import sys
import time
import uuid
from pathlib import Path

import pandas as pd
import requests

_REPO = Path(__file__).resolve().parents[1]
_LIB = _REPO / "library-onboarding"
for p in (str(_REPO), str(_LIB)):
    if p not in sys.path:
        sys.path.insert(0, p)
try:
    from dotenv import load_dotenv
    load_dotenv(_LIB / ".env", override=True)
except Exception:  # pragma: no cover
    pass

import ingest        # noqa: E402
import register      # noqa: E402
import snow          # noqa: E402
import _bulk_load_utils as bulk  # noqa: E402
from config import settings  # noqa: E402

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) ripple-data-pipeline"}
BASE = "https://www.mentalhealth.va.gov/MENTALHEALTH/docs/data-sheets/2025/"
URLS = {
    "natl": BASE + "National_Suicide_Data_Appendix_2021-2023_508.xlsx",
    "state": BASE + "State_Data_Appendix_2021-2023_508.xlsx",
    "allcause": BASE + "All-Cause_Mortality_Data_Appendix_2018-2023_508.xlsx",
}
SKIP_SHEETS = {"Note", "Figures and Tables", "Territories"}


def _clean_header(cells) -> list[str]:
    out = []
    for j, c in enumerate(cells):
        name = re.sub(r"\s+", " ", str(c).strip()) if c is not None and str(c).strip() else f"COL_{j}"
        out.append(name)
    return out


def _tidy_sheet(ws, sheet: str, extra: dict) -> pd.DataFrame:
    """Suicide-file sheets: rows 1-2 are banner/blank, row 3 is the header."""
    rows = list(ws.iter_rows(values_only=True))
    header = _clean_header(rows[2])
    data = [r for r in rows[3:] if any(c is not None and str(c).strip() for c in r)]
    df = pd.DataFrame(data, columns=header)
    for k, v in extra.items():
        df[k] = v
    return df


def _year_blocks(ws, sheet: str) -> pd.DataFrame:
    """All-cause sheets: 8-col year blocks side by side; year stamp on row 3,
    block header on row 4, data below. Unpivot to long."""
    rows = list(ws.iter_rows(values_only=True))
    year_row, header_row = rows[2], rows[3]
    ncol = len(header_row)
    frames = []
    for start in range(0, ncol, 8):
        year = year_row[start]
        if year is None or not str(year).strip():
            continue
        header = _clean_header(header_row[start:start + 8])
        data = []
        for r in rows[4:]:
            block = r[start:start + 8]
            if any(c is not None and str(c).strip() for c in block):
                data.append(block)
        blk = pd.DataFrame(data, columns=header)
        blk["YEAR"] = str(year).strip()
        frames.append(blk)
    df = pd.concat(frames, ignore_index=True)
    # cohort/sex come from the sheet name: 'Recent-VHA Veteran-Female' etc.
    m = re.match(r"(.*)-(All|Female|Male)$", sheet)
    df["COHORT"] = (m.group(1) if m else sheet).strip()
    df["SEX"] = (m.group(2) if m else "All").strip()
    return df


def _fetch(url: str) -> bytes:
    r = requests.get(url, headers=UA, timeout=120)
    r.raise_for_status()
    if not r.content[:2] == b"PK":
        raise RuntimeError(f"not an xlsx (got {r.content[:40]!r}): {url}")
    return r.content


def build_frames() -> dict[str, tuple[pd.DataFrame, str, str]]:
    """source_id -> (df, source_url, sha256 of the raw file(s))."""
    import openpyxl
    out = {}

    raw = _fetch(URLS["natl"])
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    frames = [_tidy_sheet(wb[s], s, {"COHORT": s}) for s in wb.sheetnames if s not in SKIP_SHEETS]
    out["fed_va_suicide_national"] = (
        pd.concat(frames, ignore_index=True), URLS["natl"],
        hashlib.sha256(raw).hexdigest())

    raw = _fetch(URLS["state"])
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    frames = [_tidy_sheet(wb[s], s, {"SHEET": s}) for s in wb.sheetnames if s not in SKIP_SHEETS]
    out["fed_va_suicide_state"] = (
        pd.concat(frames, ignore_index=True), URLS["state"],
        hashlib.sha256(raw).hexdigest())

    raw = _fetch(URLS["allcause"])
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    frames = [_year_blocks(wb[s], s) for s in wb.sheetnames if s not in SKIP_SHEETS]
    out["fed_va_allcause_mortality"] = (
        pd.concat(frames, ignore_index=True), URLS["allcause"],
        hashlib.sha256(raw).hexdigest())
    return out


REGISTRY = {
    "fed_va_suicide_national": {
        "name": "VA National Suicide Data Appendix (2001-2023)",
        "unit_of_observation": "one row = one cohort x year of veteran/US-adult suicide counts and rates",
        "description": "Veteran, VHA-user, other-veteran, non-veteran and US-adult suicide "
                       "deaths, populations and rates by year and sex, 2001-2023.",
        "expected_min_rows": 500,
    },
    "fed_va_suicide_state": {
        "name": "VA State Suicide Data Appendix (2001-2023)",
        "unit_of_observation": "one row = one state x year (x sex/age/method) veteran suicide record",
        "description": "Veteran suicide deaths and rates by state, sex, age group and "
                       "method, 2001-2023 (state sheets of the VA suicide data appendix).",
        "expected_min_rows": 8000,
    },
    "fed_va_allcause_mortality": {
        "name": "VA All-Cause Mortality Data Appendix (2018-2023)",
        "unit_of_observation": "one row = one cohort x sex x year x ranked cause of death",
        "description": "Leading causes of veteran death with counts, rates and years of "
                       "potential life lost, by cohort and sex, 2018-2023.",
        "expected_min_rows": 1000,
    },
}


def _register_source(conn, sid: str, url: str, rows: int) -> None:
    meta = REGISTRY[sid]
    cfg = {
        "source_id": sid, "name": meta["name"],
        "publisher": "US Department of Veterans Affairs — Office of Mental Health",
        "url": "https://www.mentalhealth.va.gov/suicide_prevention/data.asp",
        "description": meta["description"],
        "jurisdiction": "US", "category": "Health", "subcategory": "Veteran Mortality",
        "unit_of_observation": meta["unit_of_observation"],
        "geographic_scope": "United States", "access_method": "bulk", "format": "xlsx",
        "auth": {"type": "none"}, "cost": "free", "update_cadence": "annual",
        "volume": f"{rows:,} rows", "license_terms": "Public domain (US Gov)",
        "join_keys": "STATE, YEAR",
        "accountability_relevance": "Who gets hurt, counted: veteran suicide and mortality "
                                    "by state/cohort — the harm ledger for any veteran-care pattern.",
        "priority_tier": "1", "landing_table": sid.upper(),
        "notes": f"Loaded by scripts/va_mortality_load.py from {url} (2026-08-09 rebuild of "
                 "the dead PDF scrapes; LLM-free).",
    }
    snow.execute(conn, *register._merge_sql(register._build_row(cfg, {})))


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="VA mortality xlsx loader")
    ap.add_argument("--run", action="store_true")
    args = ap.parse_args(argv)

    frames = build_frames()
    for sid, (df, url, _sha) in frames.items():
        print(f"{sid}: {len(df):,} rows x {len(df.columns)} cols  "
              f"(sample cols: {list(df.columns)[:6]})")
    if not args.run:
        print("\nPREVIEW only — add --run to land.")
        return 0

    started = ingest._utcnow()
    conn = snow.connect()
    failed = []
    try:
        for sid, (df, url, sha) in frames.items():
            table = sid.upper()
            run_id = str(uuid.uuid4())
            out = ingest._stringify(df)
            out.columns = [ingest._sf_col(c) for c in out.columns]
            out = out.loc[:, ~out.columns.duplicated()]
            out[ingest.META_INGESTED_AT] = started.replace(tzinfo=None)
            out[ingest.META_SOURCE_RUN_ID] = run_id
            out[ingest.META_SRC_SHA256] = sha
            from snowflake.connector.pandas_tools import write_pandas
            ok, _c, _r, _ = write_pandas(
                conn, out, table_name=table, database=settings.raw_database,
                schema=settings.raw_schema, auto_create_table=True,
                overwrite=True, quote_identifiers=False)
            if not ok:
                raise RuntimeError(f"write_pandas failed for {table}")
            passed, report = bulk.run_quality_gate(
                conn, sid, table, run_id, row_count=len(out),
                source_url=url, expected_min_rows=REGISTRY[sid]["expected_min_rows"])
            if not passed:
                print(f"QUALITY GATE FAILED {table}: {report}")
                failed.append(sid)
                continue
            _register_source(conn, sid, url, len(out))
            print(f"LOADED {len(out):,} rows -> LIBRARY_RAW.LANDING.{table}")
    finally:
        conn.close()
    if failed:
        raise SystemExit(f"quality gate failed for: {', '.join(failed)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
