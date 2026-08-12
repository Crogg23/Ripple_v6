"""Census grid — step 3: one Excel workbook, front-loaded for reading on a
flight. Sheets in story order: read-me, the wall chart, the build roadmap,
the thing-list, the slot vocabulary, families, the sources ledger, then the
full machine layers (every table, every parked branch) with filters on.

Reads the CSVs emitted by build_grid.py. Deterministic; run any time.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
SRC = REPO / "reports" / "census_grid_2026-08-12"
OUT = SRC / "census_grid.xlsx"

# ---- palette (soft, Excel-native feel) ----
GREEN = PatternFill("solid", start_color="C6EFCE")
LIGHTGREEN = PatternFill("solid", start_color="E2EFDA")
AMBER = PatternFill("solid", start_color="FFEB9C")
RED = PatternFill("solid", start_color="FFC7CE")
GREY = PatternFill("solid", start_color="F2F2F2")
HEADER = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=16)
H2_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

CLASS_PLAIN = {
    "noun": "a thing that persists",
    "event": "something that happened",
    "event+noun": "an event with its own life",
    "link": "a connection between things",
    "code": "a vocabulary",
    "aggregate": "pre-added statistics",
    "unclassified": "unknown — flagged",
}

BRANCH_MEANING = {
    "per person served": ("Divide anything by how many humans are exposed. The single most-wanted view on the chart.",
                          "population columns exist elsewhere — needs joins"),
    "was there an inspection first (real lineage)": ("Did a look-see actually precede the finding? Only provable with row-level links.",
                                                     "needs proof a violation points at its inspection"),
    "did anyone get hurt (harm join)": ("Connect paperwork events to injuries and deaths.",
                                        "harm lives in different tables — needs joins"),
    "any trend over time": ("These tables have no date at all — they are frozen snapshots.",
                            "needs history downloads or a dated re-pull"),
    "join to the entity spine": ("These tables aren't wired to the who-is-who backbone yet.",
                                 "needs spine decisions per source"),
    "assessed vs actually collected": ("A fine on paper vs money that actually moved.",
                                       "collection outcomes live elsewhere if anywhere"),
    "events per noun via hard ID": ("Rates per company/facility need a real ID; these events only carry names.",
                                    "needs name-to-ID matching"),
    "row-level version of this aggregate": ("Publisher only gave us pre-added totals; the raw rows exist at the source.",
                                            "needs a better feed from the publisher"),
    "multi-year trending": ("We hold one year; the publisher has decades sitting free.",
                            "needs history downloads"),
    "full column inventory": ("A few tables' columns are only knowable from the warehouse itself.",
                              "needs one cheap metadata pull"),
}


def read(name: str) -> list[dict]:
    with (SRC / name).open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def style_header_row(ws, ncols: int, row: int = 1) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def sheet_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    lines: list[tuple[str, str]] = [
        ("t", "The Census Grid — the whole warehouse on one wall chart"),
        ("", ""),
        ("p", "Built 2026-08-12 from table metadata only. Zero warehouse queries. Nothing sampled, nothing skipped:"),
        ("p", "all 1,765 modeled tables are on the chart, and everything unknowable is marked, not hidden."),
        ("", ""),
        ("h", "The idea in one breath"),
        ("p", "Every table in the warehouse holds one of four kinds of rows:"),
        ("p", "   • a THING that persists (a facility, a company, a nurse, a pension plan)  → always a denominator"),
        ("p", "   • something that HAPPENED to a thing, with a date (a violation, a payment) → always a numerator"),
        ("p", "   • a CONNECTION between two things (who owns what, who works where)        → the roads"),
        ("p", "   • a VOCABULARY (violation codes, industry codes)                          → the slicers"),
        ("p", "Every measure the platform will ever compute is: happenings per thing, sliced by vocabulary, along roads."),
        ("", ""),
        ("h", "How to read the WALL CHART sheet (the main event)"),
        ("p", "Rows = the 38 families of things.  Columns = the ~40 ways of looking at anything."),
        ("p", "Each cell says how many of that family's tables can already produce that view — '12/14' means"),
        ("p", "12 of 14 tables have the columns for it. Colors:"),
        ("p", "   GREEN = every table can do it     LIGHT GREEN = most can"),
        ("p", "   AMBER = some can                  RED = none can (a structural hole — the data physically lacks it)"),
        ("p", "   grey/blank = the question doesn't apply to that kind of row"),
        ("p", "No cell holds real numbers yet. This chart is the FRAME — it proves what is answerable."),
        ("p", "Filling the cells is the next build step (needs a small warehouse spend, price tag first)."),
        ("", ""),
        ("h", "How to read the BUILD ROADMAP sheet (the second prize)"),
        ("p", "While charting, every idea that needed a second table, a join, or missing data got PARKED with a tally mark."),
        ("p", "4,357 parks collapsed into 10 branches, ranked by votes. That ranking — not anyone's hunch —"),
        ("p", "is the build order. 'Per person served' won by a landslide: 1,426 tables want it."),
        ("", ""),
        ("h", "The other sheets"),
        ("p", "THINGS — the 278 distinct kinds of things, grouped into families, with table counts."),
        ("p", "WAYS OF LOOKING — the ~40 column-headers of the wall chart, defined in plain words."),
        ("p", "FAMILIES — one line per family: what it is, how many tables, ready vs holes."),
        ("p", "SOURCES LEDGER — every source the loaders ever attempted, with status. Read the warning at the top."),
        ("p", "EVERY TABLE — the machine layer: all 1,765 tables, how each was classified and the evidence. Filterable."),
        ("p", "PARKING LOT (FULL) — all 4,357 parked branches, one line each. Filterable."),
        ("", ""),
        ("h", "The honest residue (visible on purpose)"),
        ("p", "9 tables couldn't be identified at all. 235 were classified only by the shape of their columns."),
        ("p", "674 don't declare what one row means. 12 have columns only knowable from the warehouse."),
        ("p", "They are all ON the chart, flagged — a visible hole is coverage; a quietly dropped hole is a lie."),
        ("", ""),
        ("h", "One finding worth knowing before the flight ends"),
        ("p", "The warehouse cannot currently say how many sources it holds. The loader's logbook says 774 attempted"),
        ("p", "(684 marked failed) — yet 1,141 sources are live in the build system and 1,329 raw tables exist."),
        ("p", "Three counts, no shared key. Broken bookkeeping, not broken data — and it's parked on the roadmap."),
    ]
    r = 1
    for kind, text in lines:
        cell = ws.cell(row=r, column=1, value=text)
        if kind == "t":
            cell.font = TITLE_FONT
        elif kind == "h":
            cell.font = H2_FONT
        r += 1


def sheet_wallchart(wb: Workbook) -> None:
    fams = read("grid_families.csv")
    slots = read("slots.csv")
    slot_order = [s["slot"] for s in slots]
    slot_label = {s["slot"]: s["label"] for s in slots}
    fam_rows: dict[str, dict] = {}
    cells: dict[tuple[str, str], tuple[int, int]] = {}
    for g in fams:
        fam_rows.setdefault(g["family"], {"class": g["class"], "n": int(g["n_members"])})
        cells[(g["family"], g["slot"])] = (int(g["members_ready"]), int(g["members_hole"]))
    fam_order = sorted(fam_rows, key=lambda f: -fam_rows[f]["n"])

    ws = wb.create_sheet("WALL CHART")
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="family").font = HEADER_FONT
    ws.cell(row=1, column=2, value="what its rows are").font = HEADER_FONT
    ws.cell(row=1, column=3, value="tables").font = HEADER_FONT
    for j, sid in enumerate(slot_order, start=4):
        c = ws.cell(row=1, column=j, value=slot_label[sid])
        c.alignment = Alignment(textRotation=60, wrap_text=True, vertical="bottom")
        c.fill = HEADER
        c.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(j)].width = 6.5
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = HEADER
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 7
    ws.row_dimensions[1].height = 110

    for i, fam in enumerate(fam_order, start=2):
        ws.cell(row=i, column=1, value=fam).font = BOLD
        ws.cell(row=i, column=2, value=CLASS_PLAIN.get(fam_rows[fam]["class"], fam_rows[fam]["class"]))
        ws.cell(row=i, column=3, value=fam_rows[fam]["n"]).alignment = Alignment(horizontal="center")
        for j, sid in enumerate(slot_order, start=4):
            cell = ws.cell(row=i, column=j)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
            if (fam, sid) not in cells:
                cell.fill = GREY
                continue
            ready, hole = cells[(fam, sid)]
            total = ready + hole
            cell.value = f"{ready}/{total}"
            if ready == 0:
                cell.fill = RED
            elif ready == total:
                cell.fill = GREEN
            elif ready / total >= 0.5:
                cell.fill = LIGHTGREEN
            else:
                cell.fill = AMBER
    ws.freeze_panes = "D2"


def sheet_roadmap(wb: Workbook) -> None:
    tally = read("parking_tally.csv")
    ws = wb.create_sheet("BUILD ROADMAP")
    ws.sheet_view.showGridLines = False
    headers = ["rank", "votes", "families touched", "the branch", "what it means", "what it takes"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, len(headers))
    widths = [6, 8, 10, 38, 62, 46]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, t in enumerate(tally, start=2):
        meaning, takes = BRANCH_MEANING.get(t["branch"], ("", t.get("park_type", "")))
        ws.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=int(t["times_parked"])).font = BOLD
        ws.cell(row=i, column=3, value=int(t["families_touched"])).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=t["branch"]).font = BOLD
        ws.cell(row=i, column=5, value=meaning).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=6, value=takes).alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"


def sheet_things(wb: Workbook) -> None:
    things = read("things.csv")
    ws = wb.create_sheet("THINGS")
    headers = ["family", "thing", "what its rows are", "tables", "subject areas", "example table"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, len(headers))
    for j, w in enumerate([16, 22, 24, 8, 40, 52], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    things.sort(key=lambda r: (r["family"], -int(r["n_models"])))
    for i, t in enumerate(things, start=2):
        ws.cell(row=i, column=1, value=t["family"])
        ws.cell(row=i, column=2, value=t["thing"]).font = BOLD
        ws.cell(row=i, column=3, value=CLASS_PLAIN.get(t["class"], t["class"]))
        ws.cell(row=i, column=4, value=int(t["n_models"]))
        ws.cell(row=i, column=5, value=t["subjects"])
        ws.cell(row=i, column=6, value=t["example_models"].split(";")[0])
    ws.auto_filter.ref = f"A1:F{len(things) + 1}"
    ws.freeze_panes = "A2"


def sheet_slots(wb: Workbook) -> None:
    slots = read("slots.csv")
    ws = wb.create_sheet("WAYS OF LOOKING")
    headers = ["the view", "what it needs on the table", "applies to"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, len(headers))
    for j, w in enumerate([44, 34, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    NEED = {"universal": "nothing — always possible", "date": "a date column",
            "geo": "a place column", "geo_point": "coordinates", "code": "a category column",
            "money": "a dollar column", "quantity": "a size/count column",
            "id": "a real ID column", "population": "a people-served column",
            "flag": "a yes/no column", "_key": "a declared unique key",
            "code+date": "a category column and a date"}
    for i, s in enumerate(slots, start=2):
        ws.cell(row=i, column=1, value=s["label"]).font = BOLD
        ws.cell(row=i, column=2, value=NEED.get(s["requires"], s["requires"]))
        cls = s["classes"]
        ws.cell(row=i, column=3, value="every kind of row" if cls == "all"
                else ", ".join(CLASS_PLAIN.get(c, c) for c in cls.split(",")))
    ws.freeze_panes = "A2"


def sheet_families(wb: Workbook) -> None:
    fams = read("grid_families.csv")
    agg: dict[str, dict] = {}
    for g in fams:
        a = agg.setdefault(g["family"], {"class": g["class"], "n": int(g["n_members"]), "ready": 0, "hole": 0})
        a["ready"] += int(g["members_ready"])
        a["hole"] += int(g["members_hole"])
    ws = wb.create_sheet("FAMILIES")
    headers = ["family", "what its rows are", "tables", "cells ready", "cells hole", "% ready"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, len(headers))
    for j, w in enumerate([18, 26, 8, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, (fam, a) in enumerate(sorted(agg.items(), key=lambda kv: -kv[1]["n"]), start=2):
        ws.cell(row=i, column=1, value=fam).font = BOLD
        ws.cell(row=i, column=2, value=CLASS_PLAIN.get(a["class"], a["class"]))
        ws.cell(row=i, column=3, value=a["n"])
        ws.cell(row=i, column=4, value=a["ready"])
        ws.cell(row=i, column=5, value=a["hole"])
        pct = a["ready"] / max(1, a["ready"] + a["hole"])
        c = ws.cell(row=i, column=6, value=round(pct, 2))
        c.number_format = "0%"
        c.fill = GREEN if pct >= 0.7 else (AMBER if pct >= 0.4 else RED)
    ws.freeze_panes = "A2"


def sheet_sources(wb: Workbook) -> None:
    rows = read("sources_census.csv")
    ws = wb.create_sheet("SOURCES LEDGER")
    warn = ws.cell(row=1, column=1, value="WARNING: this logbook does not reconcile with reality — it says 774 attempted / 684 failed, "
                   "yet 1,141 sources are live in the build system. Treat 'failed' as 'the LOGBOOK thinks it failed'. "
                   "Fixing this bookkeeping is on the roadmap.")
    warn.font = Font(bold=True, color="9C0006")
    warn.alignment = Alignment(wrap_text=True)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 45
    headers = ["source", "status", "attempts", "last touched"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=2, column=j, value=h)
    style_header_row(ws, len(headers), row=2)
    for j, w in enumerate([58, 14, 10, 24], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r["source"])
        st = ws.cell(row=i, column=2, value=r["status"])
        st.fill = GREEN if r["status"] == "complete" else (AMBER if r["status"] == "needs_key" else RED)
        ws.cell(row=i, column=3, value=r["attempts"])
        ws.cell(row=i, column=4, value=r["updated_at"][:19].replace("T", " "))
    ws.auto_filter.ref = f"A2:D{len(rows) + 2}"
    ws.freeze_panes = "A3"


def sheet_tables(wb: Workbook) -> None:
    rows = read("table_map.csv")
    ws = wb.create_sheet("EVERY TABLE")
    headers = ["table (model name)", "layer", "subject", "family", "kind of rows", "thing",
               "confidence", "how it was classified", "columns", "has date", "has $", "has place",
               "has ID", "has people-served", "one row is..."]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, len(headers))
    for j, w in enumerate([52, 10, 18, 14, 22, 16, 11, 46, 9, 8, 6, 9, 7, 14, 40], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    conf_fill = {"high": GREEN, "medium": LIGHTGREEN, "low": AMBER, "none": RED}
    for i, r in enumerate(rows, start=2):
        yn = lambda k: "Y" if int(r[k]) > 0 else ""
        vals = [r["model"], r["layer"], r["subject"], r["family"],
                CLASS_PLAIN.get(r["class"], r["class"]), r["thing_token"],
                r["map_confidence"], r["map_evidence"], int(r["n_columns"]),
                yn("sem_date"), yn("sem_money"), yn("sem_geo"), yn("sem_id"),
                yn("sem_population"), r["grain_phrase"]]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v)
        ws.cell(row=i, column=7).fill = conf_fill.get(r["map_confidence"], GREY)
    ws.auto_filter.ref = f"A1:O{len(rows) + 1}"
    ws.freeze_panes = "B2"


def sheet_parkinglot(wb: Workbook) -> None:
    rows = read("parking_lot.csv")
    ws = wb.create_sheet("PARKING LOT (FULL)")
    headers = ["branch", "type", "family", "table", "the one-line park"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, len(headers))
    for j, w in enumerate([38, 22, 14, 46, 70], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["branch"])
        ws.cell(row=i, column=2, value=r["park_type"])
        ws.cell(row=i, column=3, value=r["family"])
        ws.cell(row=i, column=4, value=r["model"])
        ws.cell(row=i, column=5, value=r["line"])
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"
    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    sheet_readme(wb)
    sheet_wallchart(wb)
    sheet_roadmap(wb)
    sheet_things(wb)
    sheet_slots(wb)
    sheet_families(wb)
    sheet_sources(wb)
    sheet_tables(wb)
    sheet_parkinglot(wb)
    wb.save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size/1024:.0f} KB, {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
