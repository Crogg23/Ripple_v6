"""Census grid — step 4: the EXPLAINER workbook. The reference workbook
(census_grid.xlsx) answers "what is where"; this one teaches the idea from
zero, one sheet = one idea, walked with concrete examples. Built for reading
on a flight with no one to ask questions to.

Reads the same CSVs as make_workbook.py. Illustrative numbers are labelled.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
SRC = REPO / "reports" / "census_grid_2026-08-12"
OUT = SRC / "census_grid_explainer.xlsx"

# palette
NAVY = "1F4E79"
GREEN = PatternFill("solid", start_color="C6EFCE")
LIGHTGREEN = PatternFill("solid", start_color="E2EFDA")
AMBER = PatternFill("solid", start_color="FFEB9C")
RED = PatternFill("solid", start_color="FFC7CE")
GREY = PatternFill("solid", start_color="F2F2F2")
BLUEBOX = PatternFill("solid", start_color="DDEBF7")
HEADFILL = PatternFill("solid", start_color=NAVY)
WHITE_B = Font(color="FFFFFF", bold=True, size=12)
TITLE = Font(bold=True, size=20, color=NAVY)
H2 = Font(bold=True, size=14, color=NAVY)
BODY = Font(size=12)
BODY_B = Font(size=12, bold=True)
SMALL_I = Font(size=10, italic=True, color="808080")
PUNCH = Font(bold=True, size=13, color="9C0006")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def read(name: str) -> list[dict]:
    with (SRC / name).open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def new_sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def put(ws, r, c, text, font=BODY, fill=None, wrap=True, span=None, height=None):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    if fill:
        cell.fill = fill
    if span:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
        if fill:
            for cc in range(c, c + span):
                ws.cell(row=r, column=cc).fill = fill
    if height:
        ws.row_dimensions[r].height = height
    return r + 1


def minitable(ws, r, c, headers, rows, note=None):
    """Draw a small bordered table starting at (r, c). Returns next free row."""
    for j, h in enumerate(headers):
        cell = ws.cell(row=r, column=c + j, value=h)
        cell.fill = HEADFILL
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(rows, start=1):
        for j, v in enumerate(row):
            cell = ws.cell(row=r + i, column=c + j, value=v)
            cell.border = THIN
            cell.font = Font(size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    nr = r + len(rows) + 1
    if note:
        ws.cell(row=nr, column=c, value=note).font = SMALL_I
        nr += 1
    return nr + 1


# ---------------------------------------------------------------- sheets

def s1_start(wb):
    ws = new_sheet(wb, "1 · START HERE", [3, 100])
    r = 2
    r = put(ws, r, 2, "The Census — what this is, in plain words", TITLE); r += 1
    r = put(ws, r, 2, "Your own words, and the reason all of this exists:", BODY); r += 0
    r = put(ws, r, 2, '"Dipping a 5-gallon bucket in the ocean and then saying we didn\'t catch a whale — must be nothing out there."',
            Font(size=13, italic=True, bold=True), BLUEBOX, height=30); r += 1
    r = put(ws, r, 2, "That's what pulling one story at a time out of the warehouse is. This workbook explains the alternative.", BODY); r += 1
    r = put(ws, r, 2, "The problem, in one number", H2)
    r = put(ws, r, 2, "We found that 82% of chronic violators face no enforcement. Is that the worst number in American "
            "regulation, or completely normal? NOBODY KNOWS — not us, not anyone — because no baseline exists to "
            "hold it against. A finding without a baseline is a bucket of water. You can't see the whale.", BODY, height=48); r += 1
    r = put(ws, r, 2, "The fix", H2)
    r = put(ws, r, 2, "Describe EVERYTHING in the warehouse the same way, in one language, before chasing any more "
            "stories. Then ordinary has a shape — and anything extraordinary announces itself, instead of waiting "
            "for someone to have a hunch about it.", BODY, height=48); r += 1
    r = put(ws, r, 2, "What just got built (this week, for $0 of warehouse compute)", H2)
    r = put(ws, r, 2, "•  All 1,765 tables in the warehouse were read and sorted into one grammar of four words (next sheet).", BODY)
    r = put(ws, r, 2, "•  A WALL CHART of everything × every way of looking at it — with what's answerable and what's a hole (sheets 3–4).", BODY)
    r = put(ws, r, 2, "•  A BUILD ROADMAP that ranked itself by arithmetic instead of instinct (sheet 5).", BODY); r += 1
    r = put(ws, r, 2, "How to read this workbook", H2)
    r = put(ws, r, 2, "Eight sheets, one idea each, about 15 minutes total. Sheets 2 and 3 are the whole trick — "
            "if those two land, everything else on this platform is just those two ideas repeated.", BODY, height=32)
    r = put(ws, r, 2, "The other workbook (census_grid.xlsx, same folder) is the full reference — every table, every "
            "cell, filterable. This one is the tour; that one is the shelf.", SMALL_I)


def s2_fourwords(wb):
    ws = new_sheet(wb, "2 · THE FOUR WORDS", [3, 26, 22, 20, 20, 20])
    r = 2
    r = put(ws, r, 2, "Every row in the warehouse is one of four things", TITLE, span=5); r += 1
    r = put(ws, r, 2, "All 1,765 tables. No exceptions. Once you see it, you can't unsee it.", BODY, span=5); r += 1

    r = put(ws, r, 2, "WORD 1 — A NOUN: a thing that persists. You can count them today.", H2, span=5)
    r = minitable(ws, r, 2, ["mine ID", "name", "state", "opened"],
                  [["4601234", "Marsh Fork Mine", "WV", "1998"],
                   ["1509876", "Black Thunder", "WY", "1977"],
                   ["3304567", "Bailey Mine", "PA", "1984"]],
                  "made-up rows, real shape — this is what every 'mines' table looks like")
    r = put(ws, r, 2, "Nouns are always the BOTTOM of a fraction. 'Per mine.' 'Per nursing home.' 'Per pension plan.'", BODY_B, span=5); r += 1

    r = put(ws, r, 2, "WORD 2 — AN EVENT: something that happened, on a date, to a noun.", H2, span=5)
    r = minitable(ws, r, 2, ["date", "mine ID", "what", "penalty"],
                  [["2024-03-12", "4601234", "ventilation violation", "$4,300"],
                   ["2024-03-12", "4601234", "coal dust violation", "$12,900"],
                   ["2024-06-02", "1509876", "roof support violation", "$800"]],
                  "events point at a noun (see the mine ID?) — that pointer is everything")
    r = put(ws, r, 2, "Events are always the TOP of a fraction. Violations PER MINE. Injuries PER MINE. Dollars PER MINE.", BODY_B, span=5); r += 1

    r = put(ws, r, 2, "WORD 3 — A LINK: a row that IS a relationship between two nouns.", H2, span=5)
    r = minitable(ws, r, 2, ["company", "controls", "mine ID"],
                  [["Alpha Metallurgical", "controls", "4601234"],
                   ["Alpha Metallurgical", "controls", "3304567"]],
                  "the rarest and most valuable rows in the whole warehouse — they are the roads between tables")
    r = put(ws, r, 2, "Links let a fraction TRAVEL: violations per mine → per owner → per parent company → across industries.", BODY_B, span=5); r += 1

    r = put(ws, r, 2, "WORD 4 — A CODE: a vocabulary. Never counted — used to slice.", H2, span=5)
    r = minitable(ws, r, 2, ["code", "means"],
                  [["75.400", "accumulation of combustible materials"],
                   ["75.220", "roof control plan violation"]],
                  "every 'by type' cut on any chart is a code doing its job")
    r += 1
    r = put(ws, r, 2, "THE WHOLE PLATFORM IN ONE SENTENCE:  every measure is EVENTS per NOUN, sliced by CODES, traveling along LINKS.",
            PUNCH, BLUEBOX, span=5, height=32); r += 1
    r = put(ws, r, 2, "You already run this grammar at your day job: patients are nouns, visits and med orders are events, "
            "attending-of-record is a link, diagnosis codes are codes. Same machine. Different ocean.", BODY, span=5, height=32)


def s3_onerow(wb):
    ws = new_sheet(wb, "3 · ONE ROW, WALKED", [3, 34, 30, 42])
    r = 2
    r = put(ws, r, 2, "Walk one row of the wall chart: MINES", TITLE, span=3); r += 1
    r = put(ws, r, 2, "Take one family of nouns and ask every question the chart asks. Numbers below are made up to show "
            "the shape — filling in the real ones is the next build step.", BODY, span=3); r += 1
    rows = [
        ("How many are there?", "one number", "≈ 13,000 active mines"),
        ("Over time?", "a line", "how many opened, closed, each year since the data starts"),
        ("By state?", "a ranked list", "WV, KY, PA at the top — is that just where mines are, or something else?"),
        ("On a map?", "dots", "every mine has coordinates — this one is free"),
        ("Mix by type?", "a pie", "coal vs metal vs sand&gravel — the vocabulary does the slicing"),
        ("Violations per mine?", "THE core fraction", "events ÷ nouns. The whole grammar in one cell"),
        ("Repeat offenders?", "a leaderboard", "which 1% of mines eat 40% of all violations?"),
        ("Never inspected?", "a zero-count", "mines with NO look-see events at all — absence is a finding"),
        ("Dollars?", "totals + typical", "penalty dollars: total, median, and who pays the most"),
    ]
    for q, shape, ex in rows:
        ws.cell(row=r, column=2, value=q).font = BODY_B
        ws.cell(row=r, column=3, value=shape).font = BODY
        ws.cell(row=r, column=4, value=ex).font = BODY
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=2).fill = LIGHTGREEN
        r += 1
    r += 1
    r = put(ws, r, 2, "And then the chart hits two questions the data CANNOT answer yet:", H2, span=3)
    for q, why in [("Violations per MINER?", "the violations table doesn't say how many people work at the mine — "
                    "that number lives in a different table. → PARKED, one tally mark."),
                   ("Did being fined change anything?", "needs linking a fine to what happened at that mine afterwards "
                    "— a join we haven't proven yet. → PARKED, one tally mark.")]:
        ws.cell(row=r, column=2, value=q).font = BODY_B
        ws.cell(row=r, column=2).fill = RED
        ws.cell(row=r, column=2).border = THIN
        cell = ws.cell(row=r, column=3, value=why)
        cell.font = BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 34
        r += 1
    r += 1
    r = put(ws, r, 2, "A red cell is not a failure — it's COVERAGE. We now know exactly what the data can't say, "
            "and every red cell filed a vote for what to build next (sheet 5).", PUNCH, BLUEBOX, span=3, height=32)
    r = put(ws, r, 2, "Now imagine this walk done 278 times — every kind of thing in the warehouse. That's sheet 4.", BODY, span=3)


def s4_minichart(wb):
    fams = read("grid_families.csv")
    cells = {}
    fam_n = {}
    for g in fams:
        cells[(g["family"], g["slot"])] = (int(g["members_ready"]), int(g["members_hole"]))
        fam_n[g["family"]] = int(g["n_members"])
    pick_fams = ["facility", "organization", "provider", "person", "violation",
                 "inspection", "enforcement", "accident", "payment", "contribution",
                 "case", "ownership"]
    pick_slots = [("how_many", "How many"), ("by_time", "Over time"), ("by_state", "By state"),
                  ("map_points", "On a map"), ("mix_by_code", "Mix by type"), ("total_money", "Dollars"),
                  ("per_noun_rate", "Per noun"), ("repeaters", "Repeat offenders"), ("per_person", "Per person served")]
    ws = new_sheet(wb, "4 · THE WALL CHART", [3, 16, 8] + [11] * len(pick_slots))
    r = 2
    r = put(ws, r, 2, "The wall chart — starter version, real numbers", TITLE, span=10); r += 1
    r = put(ws, r, 2, "12 of the 38 families × 9 of the 40 questions. Each cell: how many of that family's tables can "
            "already produce that view. GREEN = all · LIGHT GREEN = most · AMBER = some · RED = none (a hole) · grey = doesn't apply.",
            BODY, span=10, height=34); r += 1
    hdr = r
    ws.cell(row=hdr, column=2, value="family").font = WHITE_B
    ws.cell(row=hdr, column=2).fill = HEADFILL
    ws.cell(row=hdr, column=3, value="tables").font = WHITE_B
    ws.cell(row=hdr, column=3).fill = HEADFILL
    for j, (_sid, lab) in enumerate(pick_slots, start=4):
        c = ws.cell(row=hdr, column=j, value=lab)
        c.fill = HEADFILL
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[hdr].height = 30
    r = hdr + 1
    for fam in pick_fams:
        if fam not in fam_n:
            continue
        ws.cell(row=r, column=2, value=fam).font = BODY_B
        ws.cell(row=r, column=3, value=fam_n[fam]).alignment = Alignment(horizontal="center")
        for j, (sid, _lab) in enumerate(pick_slots, start=4):
            cell = ws.cell(row=r, column=j)
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
            if (fam, sid) not in cells:
                cell.fill = GREY
                continue
            ready, hole = cells[(fam, sid)]
            total = ready + hole
            cell.value = f"{ready}/{total}"
            cell.fill = (RED if ready == 0 else GREEN if ready == total
                         else LIGHTGREEN if ready / total >= 0.5 else AMBER)
        r += 1
    r += 1
    r = put(ws, r, 2, "Read one row out loud and it works: 'Violations — every table can count, trend, and slice them; "
            "most know the state; NONE can say per-person-served.' That red column on the right edge?", BODY, span=10, height=30)
    r = put(ws, r, 2, "It's red almost everywhere. Which is exactly why it's #1 on the roadmap — next sheet.", PUNCH, span=10)
    r = put(ws, r, 2, "The FULL chart (38 families × 40 questions, every cell) lives in the reference workbook, WALL CHART tab.", SMALL_I, span=10)


def s5_parking(wb):
    ws = new_sheet(wb, "5 · THE PARKING LOT", [3, 44, 12, 60])
    r = 2
    r = put(ws, r, 2, "The parking lot — how the roadmap wrote itself", TITLE, span=3); r += 1
    r = put(ws, r, 2, "The rule while building the chart: the moment an idea needs a SECOND table, a join, or missing data "
            "— stop. Don't chase it. Write one line, add a tally mark, get back to the surface.", BODY, span=3, height=32)
    r = put(ws, r, 2, "Why so strict? Because chasing the first shiny thread is how every previous session ended up "
            "small. Cover everything first; let the tally marks pile up; then read the pile.", BODY, span=3, height=32); r += 1
    r = put(ws, r, 2, "4,357 tally marks later, the pile ranked itself:", H2, span=3)
    data = [
        ("1,426 votes", "Per person served", "Everything wants dividing by humans exposed — violations per resident, "
         "not per nursing home. The single most-wanted thing on the platform, by arithmetic."),
        ("658 votes", "Was there an inspection first?", "Connect findings to the look-see that produced them. Without it, "
         "'82% unenforced' stays a bucket number."),
        ("658 votes", "Did anyone get hurt?", "Join paperwork events to injuries and deaths. The mission question, as a build item."),
        ("616 votes", "No date at all", "616 tables are photographs, not movies — they can never trend until history gets loaded."),
        ("615 votes", "Not wired to the spine", "615 tables aren't connected to the who-is-who backbone — their fractions can't travel yet."),
        ("160 votes", "Fined vs actually paid", "A penalty on paper vs money that moved. Famous chasm; now countable."),
    ]
    for votes, branch, meaning in data:
        ws.cell(row=r, column=2, value=branch).font = Font(bold=True, size=13)
        v = ws.cell(row=r, column=3, value=votes)
        v.font = Font(bold=True, size=13, color=NAVY)
        m = ws.cell(row=r, column=4, value=meaning)
        m.font = BODY
        m.alignment = Alignment(wrap_text=True, vertical="top")
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).border = THIN
        ws.cell(row=r, column=2).fill = BLUEBOX
        ws.row_dimensions[r].height = 44
        r += 1
    r += 1
    r = put(ws, r, 2, "NOBODY CHOSE THIS RANKING. It emerged from describing everything and counting what got parked. "
            "That's instinct replaced by arithmetic — the whole point of the census.", PUNCH, BLUEBOX, span=3, height=32)


def s6_ratios(wb):
    ws = new_sheet(wb, "6 · WHERE THIS GOES", [3, 20, 15, 15, 15, 15, 15, 15])
    r = 2
    r = put(ws, r, 2, "The destination: one page nobody else on earth can make", TITLE, span=7); r += 1
    r = put(ws, r, 2, "When the grid fills, every domain answers the SAME six questions in the SAME units, side by side. "
            "No newsroom, agency, or watchdog holds every regulator's record on one spine. This warehouse does.", BODY, span=7, height=32); r += 1
    heads = ["", "How often does anyone look?", "How often is it wrong when they look?", "Does getting caught mean anything?",
             "Does it cost anything?", "Does paperwork predict harm?", "How many people per institution?"]
    doms = ["Mines", "Nursing homes", "Pensions", "Drug makers", "Banks", "Charities", "Contractors"]
    for j, h in enumerate(heads):
        c = ws.cell(row=r, column=2 + j, value=h)
        c.fill = HEADFILL
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 44
    r += 1
    for d in doms:
        ws.cell(row=r, column=2, value=d).font = BODY_B
        ws.cell(row=r, column=2).border = THIN
        for j in range(6):
            c = ws.cell(row=r, column=3 + j, value="?")
            c.border = THIN
            c.alignment = Alignment(horizontal="center")
            c.fill = GREY
        r += 1
    r += 1
    r = put(ws, r, 2, "Every '?' becomes a number when the grid cells fill. Then the page reads like an X-ray:", BODY, span=7)
    r = put(ws, r, 2, "if every domain's funnel collapses at the same step — say, plenty of looking, plenty of findings, "
            "almost no consequences — that is a story about American regulation no one has ever been able to SHOW. "
            "And if one domain looks nothing like its neighbors, that's where the light points next.", BODY, span=7, height=44); r += 1
    r = put(ws, r, 2, "One honest rule, locked in from day one: domains compare by SHAPE (where the funnel collapses), "
            "never by LEVEL ('mining is stricter than eldercare') — agencies cite different things under different laws.", SMALL_I, span=7, height=30)


def s7_holes(wb):
    ws = new_sheet(wb, "7 · HONEST HOLES", [3, 100])
    r = 2
    r = put(ws, r, 2, "What this census admits about itself", TITLE); r += 1
    r = put(ws, r, 2, "The rule the whole build ran on: a visible hole is coverage; a quietly dropped hole is a lie. So:", BODY); r += 1
    for line in [
        "•  9 tables couldn't be identified at all — they're on the chart, flagged, not hidden.",
        "•  235 tables were classified only by the shape of their columns (educated guess, marked as such).",
        "•  674 tables never declare what one row means — their 'one row = one what?' is unstated.",
        "•  12 tables have columns only knowable by asking the warehouse itself (one cheap pull fixes it).",
        "•  No cell on the wall chart holds a real number yet — the chart is the FRAME, proven fillable.",
    ]:
        r = put(ws, r, 2, line, BODY)
    r += 1
    r = put(ws, r, 2, "And the one that matters most", H2)
    r = put(ws, r, 2, "The warehouse cannot currently say how many sources it holds. The loader's logbook says 774 "
            "attempted and 684 failed — yet 1,141 sources are live in the build system and 1,329 raw tables exist. "
            "Three counts, no shared key between them. Broken bookkeeping, not broken data — and it's on the "
            "roadmap, because 'how much do we even hold' is the denominator under everything else.", BODY, height=60)


def s8_next(wb):
    ws = new_sheet(wb, "8 · WHAT HAPPENS NEXT", [3, 100])
    r = 2
    r = put(ws, r, 2, "The build order from here", TITLE); r += 1
    steps = [
        ("STEP 1 — Fill the chart (next session)", "Run cheap scans over every table: real counts, real date ranges, "
         "real ID health. The 30,509 green cells light up with numbers. Price tag comes to you before any query runs — "
         "expect small; this is counting, not crunching."),
        ("STEP 2 — The ratios page", "Compute the six questions per domain (sheet 6). First version will have holes — "
         "the holes are themselves the map of what to wire next."),
        ("STEP 3 — Work the roadmap by vote count", "Per-person-served first (1,426 votes), then lineage, then the "
         "harm joins. Each one flips thousands of cells from red to green."),
        ("STEP 4 — Then, and only then, go hunting", "With a baseline under everything, 'is this weird?' becomes a "
         "lookup instead of a debate. The ten findings we already have become readable cells in the frame — and the "
         "next ones will announce themselves."),
    ]
    for h, b in steps:
        r = put(ws, r, 2, h, H2)
        r = put(ws, r, 2, b, BODY, height=46)
        r += 1
    r = put(ws, r, 2, "Nothing above spends real money without a price tag in your chat first. Nothing publishes without "
            "your sign-off. Same rules as always.", SMALL_I)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    s1_start(wb)
    s2_fourwords(wb)
    s3_onerow(wb)
    s4_minichart(wb)
    s5_parking(wb)
    s6_ratios(wb)
    s7_holes(wb)
    s8_next(wb)
    wb.save(OUT)
    print(f"wrote {OUT} ({OUT.stat().st_size/1024:.0f} KB, {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
